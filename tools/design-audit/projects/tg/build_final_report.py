#!/usr/bin/env python3
"""TG — build the FINAL clean Design-QC PDF (NHAPOA format) from authored findings.

Single source of truth = the SCREENS / LOGIN / PARITY / GLOBAL tables below (synced from
FINDINGS.md + the login boards). Assembles audit-master.json and renders the fixed-style
report via the docs-side generate_pdf.py into docs/qc/portals/tg/. Pinned screens use
qc_geometry.finalize (percent anchors); the GLOBAL page and parity screens set boxes explicitly.

Run: python3 build_final_report.py
"""
import json, os, sys, subprocess
ENG = os.path.dirname(os.path.abspath(__file__)); TOOL = os.path.dirname(os.path.dirname(ENG))
sys.path.insert(0, os.path.join(TOOL, "engine"))
import qc_geometry as G
PROJ = ENG; OUT = os.path.join(PROJ, "out"); CAP = os.path.join(PROJ, "captures")
DEST = "/Users/akashk/Documents/Projects/MoSJE/docs/qc/portals/tg"
FURL = "https://www.figma.com/design/gH2vQ62cfg4677YKWuOpLc/MoSJE-Portal--Handoff-?node-id={n}"
ADM = "https://tg-admin-dev.mosje.in"; CIT = "https://tg-user-dev.mosje.in"
def furl(n): return FURL.format(n=n.replace(":", "-")) if n else None
def ap(rel): return os.path.join(PROJ, rel) if rel else None

# ---- per-screen findings: [sev, axis, title, design-intent, live-build, fix, xpct, ypct] ----
SCREENS = [
 {"slug": "CAD", "role": "Central Admin", "name": "Dashboard", "node": "2494:38830",
  "figma": "captures/figma/design-ADMIN-DASH.png", "live": "captures/live/CENTRAL-ADMIN-DASHBOARD.png",
  "route": ADM + "/dashboard", "f": [
   ["Nit", "Layout & Spacing", "Global filter bar sits above the KPI cards",
    "The design keeps the State / District / Date-range filters inside the Application Queue card.",
    "The build floats a global filter bar ABOVE the KPI cards; it is a richer analytics dashboard (6 KPIs + State/District charts) vs the design's 4-KPI exception queue.",
    "Confirm the analytics layout is intended — the extra KPIs/charts are build-only (audit vs the visual language, don't remove).", 50, 15],
   ["Minor", "Color & Token", "Confirm analytics chart palette uses TG tokens",
    "Charts should use TG tokens — Primary/Source #003366 for bars, Success/Source #2e7d32 for the approval-rate series.",
    "Chart series may use raw greens/blues rather than the TG token palette.",
    "Verify the chart palette against the TG tokens; the charts are a build-only section audited vs the visual language.", 40, 62]]},
 {"slug": "DMD", "role": "District Magistrate", "name": "Dashboard", "node": "2494:42053",
  "figma": "captures/figma/design-DM-DASH.png", "live": "captures/live/DISTRICT-MAGISTRATE-DASHBOARD.png",
  "route": ADM + "/dashboard", "f": [
   ["Nit", "Components & States", "Extra 'Current Stage' column + View action",
    "The design DM queue columns are Applicant ID, Name, District, Due in, Status.",
    "The build queue adds a 'Current Stage' column and a View action not in the design table.",
    "Confirm if intended — the build has an extra column; do not remove without design sign-off.", 82, 55]]},
 {"slug": "EOD", "role": "Examining Officer (Maker)", "name": "Dashboard", "node": "2494:41744",
  "figma": "captures/figma/design-MAKER-DASH.png", "live": "captures/live/EXAMINING-OFFICER-DASHBOARD.png",
  "route": ADM + "/dashboard", "f": [
   ["Minor", "Components & States", "'Classification' column dropped",
    "The design Maker queue has a Classification column (Clean / Exception badges) flagging exception applications.",
    "The build omits the Classification column, so triage exceptions aren't visible in the list.",
    "Show the Classification column per the design (low priority — the build shows fewer columns than the design).", 62, 55],
   ["Minor", "Components & States", "Queue filters missing",
    "The design queue header carries All Classifications / All Types / All Status filter dropdowns.",
    "The build has only a search box + a State/District toggle.",
    "Surface the relevant queue filters per the design.", 40, 30]]},
 {"slug": "AD", "role": "Admin", "name": "Application Detail", "node": "2494:38975",
  "figma": "captures/figma/design-APP-DETAIL.png", "live": "captures/live/CENTRAL-ADMIN-APPLICATION-DETAIL.png",
  "route": ADM + "/applications", "f": [
   ["Minor", "Layout & Spacing", "Persistent expanded admin sidebar narrows the content",
    "The design detail page is full-width (sidebar collapsed to a hamburger); the examining-officer / DM detail pages have no sidebar (match the design).",
    "The central-admin detail page keeps the admin sidebar expanded, narrowing the content column.",
    "Confirm the central-admin sidebar on detail views is intended; otherwise collapse it to match the design.", 10, 42],
   ["Nit", "Components & States", "Detail tab set differs (2 design / 4 build)",
    "The design detail has 2 tabs: Applicant Details, Documents.",
    "The build has 4 tabs — it adds Revised Certificate and ID Card.",
    "Build-extra — confirm if intended.", 45, 22]]},
]

# ---- citizen journey screens reviewed in the per-screen pass (design synced from Figma) ----
CITIZEN = [
 {"slug": "CIT-APPLY-IDENTITY", "role": "Citizen", "name": "Apply · Step 1 · Basic Identity Details", "node": "3531:35553",
  "figma": "captures/figma/dframe-citizen-apply-identity.png", "live": "captures/live/CITIZEN-CH-APPLY-IDENTITY-FILLED.png",
  "route": CIT + "/apply", "f": [
   ["Minor", "Layout & Spacing", "Form grid drops from 3 columns to 2",
    "Self-Perceived Identity and Permanent Address lay out on a 3-column field grid.",
    "The build uses a 2-column grid, making the form noticeably taller and pushing the primary action below the fold.",
    "Match the design's 3-column field grid on the identity step.", 50, 30],
   ["Minor", "Components & States", "Correspondence-address control changed from a checkbox to Yes/No buttons",
    "A single checkbox reads 'My Current / Correspondence address is the same as my Permanent address.'",
    "The build asks 'Is your correspondence address the same as your permanent address?' with Yes/No toggle buttons.",
    "Use the design's single checkbox for the correspondence-address control.", 50, 80],
   ["Minor", "Components & States", "Bottom-left action is 'Back'; the design says 'Cancel'",
    "The bottom-left control on the step is 'Cancel'.",
    "The build renders 'Back' instead.",
    "Relabel the bottom-left control to 'Cancel' per the design.", 25, 88],
   ["Minor", "Components & States", "Active step marker is an outlined ring, not a solid disc",
    "The active step is a solid navy disc with a white numeral.",
    "The build renders the active step as an outlined ring, so it reads like an inactive step.",
    "Give the active step the design's solid navy fill + white numeral.", 18, 11]]},
 {"slug": "CIT-DASH", "role": "Citizen", "name": "Dashboard · Certificate Active", "node": "3531:36919",
  "figma": "captures/figma/dframe-citizen-dashboard-approved.png", "live": "captures/live/CITIZEN-DASHBOARD.png",
  "route": CIT + "/dashboard", "f": [
   ["Minor", "Layout & Spacing", "Welfare Benefits cards wrap to a second row",
    "All four benefit cards — Scholarships, Skill Training, Garima Greh, Medical Support — sit in a single 4-up row.",
    "The build's cards are wider, so only three fit and Medical Support wraps onto a second row.",
    "Match the design's 4-up Welfare Benefits grid so the row doesn't wrap.", 50, 60],
   ["Minor", "Components & States", "Sidebar nav has no icons and omits 'Scholarships'",
    "The sidebar lists Dashboard / Certificate/ID / Scholarships / Grievances, each with a leading icon.",
    "The build's sidebar has no per-item icons and drops the Scholarships entry (it shows a 'Transgender' scheme dropdown above the list instead).",
    "Restore the per-item sidebar icons and the Scholarships nav entry per the design.", 8, 25]]},
 {"slug": "CIT-CERT-DETAIL", "role": "Citizen", "name": "Certificate & Identity Card", "node": "3531:36666",
  "figma": "captures/figma/dframe-citizen-certificate-detail.png", "live": "captures/live/CITIZEN-CERTIFICATE-DETAIL.png",
  "route": CIT + "/certificate", "f": [
   ["Major", "Components & States", "'Other Services' section replaced by the dashboard's Welfare Benefits",
    "The certificate page carries an OTHER SERVICES section with four lifecycle actions — New Transgender Certificate & ID, Revised Certificate (Post-Medical Intervention), Correction or Update Existing Details, and Withdraw a Pending Application.",
    "The build renders the dashboard's Welfare Benefits cards here instead, so none of the certificate lifecycle actions are reachable from this page.",
    "Restore the Other Services section (New certificate · Revised certificate · Correction/Update · Withdraw) on the certificate page per the design.", 50, 72],
   ["Minor", "Typography", "Page title and intro paragraph dropped",
    "The page opens with the title 'Transgender Certificate & Identity Card' and an explanatory paragraph about the certificate.",
    "The build starts straight at the approval card — no page title or intro copy.",
    "Render the page title and intro paragraph per the design.", 30, 10],
   ["Nit", "Components & States", "Action button set differs from the design",
    "The approval card offers 'Download Certificate' and 'Download ID Card'.",
    "The build offers 'Download Revised Certificate', 'Download ID Card', 'Gender Revision Request' and 'Collapse'.",
    "Confirm the extra actions are intended and align the primary label with the design.", 35, 40]]},
 {"slug": "CIT-TRACK", "role": "Citizen", "name": "Track Status", "node": "3531:36747",
  "figma": "captures/figma/dframe-citizen-approved-track.png", "live": "captures/live/CITIZEN-CH-TRACK-APPROVED.png",
  "route": CIT + "/track-status", "f": [
   ["Major", "Components & States", "Track / status page is not built — the route lands on the dashboard",
    "A Track page shows the application progress stepper (Submitted → In Scrutiny → Approval → Certificate Generation) with the current step highlighted, plus Withdraw Application, Correction Request and View Full Application actions.",
    "The build's track route renders the citizen dashboard instead — the progress stepper and its three actions do not exist anywhere in the build.",
    "Build the track/status page per the design frame, including the progress stepper and the Withdraw / Correction Request / View Full Application actions.", 50, 38]]},
]

DM_MODALS = [
 {"slug": "DM-MODAL-APPROVE", "role": "District Magistrate", "name": "Approve modal", "node": "2494:39988",
  "figma": "captures/figma/design-APPROVE-MODAL.png", "live": "captures/live/DISTRICT-MAGISTRATE-MODAL-APPROVE.png",
  "route": ADM + "/applications", "f": [
   ["Minor", "Color & Token", "Applicant info box uses a neutral tint, not the design's success tint",
    "The applicant confirm box is green-tinted with a green border, reinforcing the approve action.",
    "The build renders a neutral grey box, losing the success affordance.",
    "Apply the design's success tint + border to the applicant confirm box.", 50, 56],
   ["Minor", "Content & Iconography", "Applicant photograph not rendered in the confirm modal",
    "The modal shows the applicant's photograph with a success check, so the approver can verify identity before committing.",
    "The build renders a generic placeholder avatar instead of the applicant's uploaded photo.",
    "Render the applicant's photograph in the approve modal; confirm whether the placeholder is only a test-data artefact.", 50, 32],
   ["Nit", "Components & States", "Build adds a 'Remarks (Optional)' field",
    "The approve modal has no remarks input.",
    "The build adds a 'Remarks (Optional)' textarea before the actions.",
    "Confirm the remarks field is intended; if so, add it to the design frame.", 50, 68]]},
]

LOGIN = [
 {"slug": "CIT-SIGNIN", "role": "Citizen", "name": "Sign In", "node": "9379:115794",
  "figma": "captures/figma/design-CITIZEN-LOGIN.png", "live": "captures/live/CITIZEN-LOGIN.png",
  "route": CIT + "/auth/sign-in", "f": [
   ["Major", "Components & States", "Role-tab switcher removed",
    "The login card leads with a three-way role switcher — Citizen · Admin · Garima Greh — with Citizen active.",
    "The citizen build renders no role tabs; the card opens straight into the 'Log In to your account' heading.",
    "Restore the Citizen / Admin / Garima Greh role-tab switcher at the top of the login card, per the design frame.", 80, 24],
   ["Major", "Components & States", "Extra Mobile Number field",
    "The design offers a single Email field under 'or sign in with credentials', then a Send OTP button.",
    "The build inserts a second 'Mobile Number' field (with its own 'or' divider) below the Email Id field — absent from the design.",
    "Align to the design's single-field pattern, or update the design frame if the Mobile Number entry is a deliberate addition.", 80, 58],
   ["Minor", "Typography", "Heading casing ('Log In' vs 'Log in')",
    "The heading reads 'Log in to your account' (sentence case).",
    "The build renders 'Log In to your account' — 'In' is capitalised.",
    "Match the design's sentence-case heading: 'Log in to your account'.", 72, 24]]},
 {"slug": "ADM-SIGNIN", "role": "Admin", "name": "Log In", "node": "9387:138143",
  "figma": "captures/figma/design-ADMIN-LOGIN.png", "live": "captures/live/ADMIN-LOGIN.png",
  "route": ADM + "/login", "f": [
   ["Major", "Components & States", "Extra Mobile Number field",
    "The design offers a single Email field, then a Send OTP button.",
    "The build adds a second 'Mobile Number' field (with an 'or' divider) below Email — absent from the design.",
    "Align to the design's single Email-field pattern, or update the design frame if Mobile Number entry is a deliberate addition.", 80, 52],
   ["Minor", "Typography", "Heading casing",
    "The heading reads 'Log in to your account' (sentence case).",
    "The build renders 'Log In to your account' — 'In' is capitalised.",
    "Match the design's sentence-case heading: 'Log in to your account'.", 72, 29]]},
]

# parity/reference screens — verified faithful (no fidelity defects); rendered as a ✓ faithful board
PARITY = [
 {"slug": "PORTAL-PICKER", "role": "All roles", "name": "Portal Picker (Change)", "node": "8103:39372",
  "figma": "captures/figma/design-LOGIN-PORTAL-PICKER.png", "live": "captures/live/CITIZEN-LOGIN-CHANGE.png",
  "route": CIT + "/auth/sign-in",
  "note": "The '⇄ Change' button in the login footer opens this portal chooser. Verified faithful to the design (same 'Choose a portal to login' list + layout)."}]

# ---- coverage boards: every remaining captured screen, shown DESIGN │ BUILD. Portal-wide Global
#      findings apply; screen-specific pins are added as the per-screen review deepens. ----
# (slug, role, name, node, figrel, liverel, route)
COVERAGE = [
 ("CIT-PREAPP", "Citizen", "Apply · Pre-application", "3531:36958", "captures/figma/dframe-citizen-apply-preapp.png", "captures/live/CITIZEN-MH-DASHBOARD-PREAPP.png", CIT + "/dashboard"),
 ("CIT-APPLY-DOCS", "Citizen", "Apply · Step 2 · Documents", "3531:35457", "captures/figma/dframe-citizen-apply-documents-filled.png", "captures/live/CITIZEN-CH-APPLY-DOCUMENTS-FILLED.png", CIT + "/apply"),
 ("CIT-APPLY-REVIEW", "Citizen", "Apply · Step 3 · Review", "3531:35334", "captures/figma/dframe-citizen-apply-review.png", "captures/live/CITIZEN-CH-APPLY-REVIEW.png", CIT + "/apply"),
 ("CIT-APPLY-CONFIRM", "Citizen", "Apply · Confirmation", "3531:35271", "captures/figma/dframe-citizen-apply-confirmation.png", "captures/live/CITIZEN-CH-APPLY-CONFIRMATION.png", CIT + "/apply"),
 ("CIT-CERT-ID", "Citizen", "Certificate · ID Card", "3531:36841", "captures/figma/dframe-citizen-certificate-id.png", "captures/live/CITIZEN-CERTIFICATE-ID.png", CIT + "/certificate"),
 ("EO-DETAIL", "Examining Officer (Maker)", "Application Detail", "2494:38975", "captures/figma/design-APP-DETAIL.png", "captures/live/EXAMINING-OFFICER-APPLICATION-DETAIL.png", ADM + "/applications"),
 ("EO-DOCS", "Examining Officer (Maker)", "Application Documents", "2494:39123", "captures/figma/design-APP-DETAIL-DOCS.png", "captures/live/EXAMINING-OFFICER-APPLICATION-DOCUMENTS.png", ADM + "/applications"),
 ("DM-DETAIL", "District Magistrate", "Application Detail · Actionable", "2494:38975", "captures/figma/design-APP-DETAIL.png", "captures/live/DISTRICT-MAGISTRATE-APPLICATION-DETAIL-ACTIONABLE.png", ADM + "/applications"),
 ("DM-DOCS", "District Magistrate", "Application Documents", "2494:39123", "captures/figma/design-APP-DETAIL-DOCS.png", "captures/live/DISTRICT-MAGISTRATE-APPLICATION-DOCUMENTS.png", ADM + "/applications"),
 ("DM-MODAL-CORRECTION", "District Magistrate", "Request-Correction modal", "2494:39801", "captures/figma/design-CORRECTION-MODAL.png", "captures/live/DISTRICT-MAGISTRATE-MODAL-CORRECTION.png", ADM + "/applications"),
 ("CA-DOCS", "Central Admin", "Application Documents", "2494:39123", "captures/figma/design-APP-DETAIL-DOCS.png", "captures/live/CENTRAL-ADMIN-APPLICATION-DOCUMENTS.png", ADM + "/applications"),
]
# build-only screens — no design frame in the audited section
# (slug, role, name, liverel, route)
BUILDONLY = [
 ("DM-MODAL-REJECT", "District Magistrate", "Reject modal (build-extra)", "captures/live/DISTRICT-MAGISTRATE-MODAL-REJECT.png", ADM + "/applications"),
 ("CA-USERS", "Central Admin", "User Management", "captures/live/CENTRAL-ADMIN-USER-MANAGEMENT.png", ADM + "/users"),
 ("CA-ROLES", "Central Admin", "Role Management", "captures/live/CENTRAL-ADMIN-ROLE-MANAGEMENT.png", ADM + "/roles"),
 ("CA-TENANTS", "Central Admin", "Tenants", "captures/live/CENTRAL-ADMIN-TENANTS.png", ADM + "/tenants"),
 ("CA-PWPOLICY", "Central Admin", "Password Policy", "captures/live/CENTRAL-ADMIN-PASSWORD-POLICY.png", ADM + "/password-policy"),
 ("CIT-GRIEVANCES", "Citizen", "Grievances", "captures/live/CITIZEN-GRIEVANCES.png", CIT + "/grievances"),
]
COVNOTE = ("Portal-wide Global findings apply to this screen (masthead co-branding, KPI cards, tables, pagination "
           "— see the Global section). Design │ build shown for review; open the Figma frame ↗ for the intended design.")
BONOTE = ("No design frame in the audited section — audited against the design-system visual language; the "
          "portal-wide Global findings apply.")
def cov_screen(t):
    slug, role, name, node, figrel, liverel, route = t
    return {"slug": slug, "name": role + " — " + name, "env": "dev", "figmaImg": ap(figrel), "liveImg": ap(liverel),
            "figmaUrl": furl(node), "liveUrl": route, "findings": [], "_role": role, "_node": node,
            "_refbadge": "coverage", "_refsub": "design │ build", "_refchip": "#1558B0", "note": COVNOTE}
def bo_screen(t):
    slug, role, name, liverel, route = t
    return {"slug": slug, "name": role + " — " + name, "env": "dev", "figmaImg": None, "liveImg": ap(liverel),
            "figmaUrl": None, "liveUrl": route, "findings": [], "_role": role,
            "_refbadge": "build-only", "_refsub": "live build", "_refchip": "#6b7280", "note": BONOTE}

# ---- GLOBAL findings (Scope: Global) — grouped by representative frame ----
GREFS = {
 "DASH": ("2494:38830", "captures/figma/design-ADMIN-DASH.png", "captures/live/CENTRAL-ADMIN-DASHBOARD.png",
          ADM + "/dashboard", "Central Admin — Dashboard"),
 "DOCS": ("2494:39123", "captures/figma/design-APP-DETAIL-DOCS.png", "captures/live/EXAMINING-OFFICER-APPLICATION-DOCUMENTS.png",
          ADM + "/applications", "Examining Officer — Application Documents"),
}
GLOBAL_GROUPS = [("DASH", "Government masthead · dashboards · tables"),
                 ("DOCS", "Application review — validation & documents")]
# (gid, sev, axis, ref, title, design-intent, live-build, fix, pinx%, piny%)
GLOBAL = [
 (1, "Major", "Content & Iconography", "DASH", "Masthead co-branding lockup removed",
  "The masthead right zone carries the SAMAVESH co-branding lockup — the Digital India logo + the SAMAVESH logo — before the user/avatar block.",
  "Both logos are omitted portal-wide; only the user name/role/avatar remain on the right. (Present on the login hero, missing on every authenticated inner page.)",
  "Restore the Digital India + SAMAVESH co-branding lockup in the masthead right zone per the design, on every screen.", 80, 5),
 (2, "Minor", "Content & Iconography", "DASH", "Masthead identity lockup reduced from 3 lines to 2",
  "'Government of India / Ministry of Social Justice & Empowerment / Department of Social Justice & Empowerment' (Department line in bold navy).",
  "Only 'Government of India / Ministry of Social Justice & Empowerment' — the Department line is dropped and the bold weight moves up a line.",
  "Render the full 3-line identity lockup with the Department line bold, per the design masthead.", 17, 6),
 (3, "Major", "Components & States", "DASH", "KPI cards lost their metric icon + trend line",
  "Each dashboard KPI card shows a leading metric icon and a delta/trend line ('▲ +14.5% vs last month' or a contextual subtitle).",
  "KPI cards render only a label + number — the icon and trend/subtext are dropped, losing the at-a-glance comparison hierarchy.",
  "Restore the KPI card icon and trend/delta subtitle per the design KPI component. Applies to all role dashboards.", 25, 24),
 (4, "Minor", "Components & States", "DASH", "Pagination active-page state differs",
  "Active page = outlined (white fill, navy border, navy numeral).",
  "Active page = a navy-filled chip with a white numeral.",
  "Match the pagination active-page control to the design's outlined style. Applies to every paginated table.", 50, 92),
 (5, "Minor", "Color & Token", "DASH", "Data-table header fill inconsistent across the build",
  "Table headers use a light neutral-gray fill (≈#f9fafb, Stroke/50).",
  "Dashboard queue tables use a near-neutral header, but the User/Role/Tenant management tables use a light blue header tint (≈#dbeafe).",
  "Standardise all data-table headers to the neutral-gray header token; drop the blue tint on the management tables.", 52, 55),
 (6, "Major", "Components & States", "DOCS", "System Validation omits the Duplicate Check row",
  "System Validation lists three checks — Data Validation, Duplicate Check (with the match + similarity score, e.g. 'Potential match: TG2024-LU-04521 (85% similarity)') and Document Verification.",
  "The build shows only Data Validation and Document Verification — the Duplicate Check row is absent, so the duplicate-detection signal never reaches the reviewer. Repeats on every role's application review screen.",
  "Restore the Duplicate Check row in System Validation with its match reference + similarity status. Applies to every role's application review screen.", 26, 42),
 (7, "Major", "Content & Iconography", "DOCS", "Documents listed by raw upload filename, not semantic label",
  "Each document is listed by its semantic name — 'Signed Affidavit', 'Aadhaar Card', 'Passport Photo' — with a matching document-type icon.",
  "The build lists raw upload filenames ('form1 (2).pdf', 'Screenshot 2026-06-24 221031.pdf', 'tgportal-image.jpg') against a generic file glyph, so a reviewer cannot tell which document is which without opening each one. Repeats on every role's Documents tab.",
  "Label each document by its semantic type with the matching type icon and keep the filename as secondary text. Applies to every role's Documents tab.", 26, 64),
 (8, "Minor", "Components & States", "DOCS", "Exception Flagged banner never rendered",
  "When the system detects a similar record the screen surfaces an amber 'Exception Flagged: Duplicate' banner above the application, explaining the match.",
  "The build never renders the exception banner, so flagged applications look identical to clean ones.",
  "Render the exception banner when the duplicate/exception check trips, per the design.", 26, 16),
]

# ---- assemble ----
def mk_findings(prefix, flist):
    o = []
    for i, f in enumerate(flist, 1):
        sev, axis, title, dz, lz, fix, xp, yp = f
        o.append({"num": i, "id": f"TG-{prefix}-{i:03d}", "element": title, "section": prefix.lower(),
                  "scope": "Screen", "axis": axis, "severity": sev, "figma": dz, "live": lz, "fix": fix,
                  "_lpct": (xp, yp), "_fpct": (xp, yp)})
    return o

def screen(s):
    return {"slug": s["slug"], "name": s["role"] + " — " + s["name"], "env": "dev",
            "figmaImg": ap(s["figma"]), "liveImg": ap(s["live"]),
            "figmaUrl": furl(s["node"]), "liveUrl": s["route"], "findings": mk_findings(s["slug"], s["f"]),
            "_role": s["role"], "_node": s["node"]}

def parity_screen(s):
    sc = {"slug": s["slug"], "name": s["role"] + " — " + s["name"], "env": "dev",
          "figmaImg": ap(s["figma"]), "liveImg": ap(s["live"]),
          "figmaUrl": furl(s["node"]), "liveUrl": s["route"], "findings": [],
          "_role": s["role"], "_node": s["node"]}
    if s.get("note"): sc["note"] = s["note"]
    return sc

def global_screens():
    byref = {}
    for (gid, sev, cat, ref, title, dz, lz, fix, px, py) in GLOBAL:
        byref.setdefault(ref, []).append((gid, sev, cat, title, dz, lz, fix, px, py))
    out = []
    for ref, label in GLOBAL_GROUPS:
        items = sorted(byref.get(ref, []), key=lambda t: t[0])
        if not items: continue
        node, dpng, bpng, route, rlabel = GREFS[ref]
        Hd = G.imgdims(dpng, PROJ)[1]; Hb = G.imgdims(bpng, PROJ)[1]
        fnds = []
        for (gid, sev, cat, title, dz, lz, fix, px, py) in items:
            fnds.append({"num": gid, "id": f"TG-GLOBAL-{gid:03d}", "element": title, "section": "global",
                         "scope": "Global", "axis": cat, "severity": sev, "figma": dz, "live": lz, "fix": fix,
                         "subO": "Scope: Global — repeats across the portal",
                         "figmaBox": [0, 0, 1440, Hd], "liveBox": [0, 0, 1440, Hb], "sectionBox": [0, 0, 1440, Hb],
                         "figmaPin": {"x": px, "y": py}, "livePin": {"x": px, "y": py}})
        out.append({"slug": f"GLOBAL-{ref}", "name": f"Global · {label}", "env": "dev",
                    "figmaImg": ap(dpng), "liveImg": ap(bpng), "figmaUrl": furl(node), "liveUrl": route,
                    "findings": fnds, "_role": "Global", "_node": node, "_refFrame": True,
                    "note": f"Scope: Global — these elements repeat across the portal; fix once. The board shows a "
                            f"representative frame ({rlabel}) — open it via 'Figma frame ↗'; each marker's number matches its TG-GLOBAL id."})
    return out

# pinned screens (admin roles + login + reviewed citizen/modals) — geometry via qc_geometry.finalize
admin_scr = [screen(s) for s in SCREENS]
login_scr = [screen(s) for s in LOGIN]
citizen_scr = [screen(s) for s in CITIZEN]
dmmodal_scr = [screen(s) for s in DM_MODALS]
for sc in admin_scr + login_scr + citizen_scr + dmmodal_scr:
    G.finalize(sc, eng_dir=CAP, base_dir=PROJ)
G.write_failures(OUT)

# report order: GLOBAL · Login · Citizen journey · Admin roles (each: pinned dash/detail + coverage) · build-only · parity
glob = global_screens()
parity_scr = [parity_screen(s) for s in PARITY]
cov = [cov_screen(t) for t in COVERAGE]
bo = [bo_screen(t) for t in BUILDONLY]
def by_role(lst, *names): return [s for s in lst if s["_role"] in names]
def by_slug(lst, *slugs): return [s for s in lst if s["slug"] in slugs]

# citizen journey in flow order (reviewed screens interleaved with coverage boards)
CIT_ORDER = ["CIT-PREAPP", "CIT-APPLY-IDENTITY", "CIT-APPLY-DOCS", "CIT-APPLY-REVIEW", "CIT-APPLY-CONFIRM",
             "CIT-DASH", "CIT-CERT-DETAIL", "CIT-CERT-ID", "CIT-TRACK"]
_cit_pool = {s["slug"]: s for s in (citizen_scr + by_role(cov, "Citizen"))}
citizen_journey = [_cit_pool[s] for s in CIT_ORDER if s in _cit_pool]
eo = by_role(admin_scr, "Examining Officer (Maker)") + by_role(cov, "Examining Officer (Maker)")
dm = by_role(admin_scr, "District Magistrate") + by_role(cov, "District Magistrate") + dmmodal_scr
ca = by_role(admin_scr, "Central Admin", "Admin") + by_role(cov, "Central Admin")
buildonly_admin = by_slug(bo, "DM-MODAL-REJECT", "CA-USERS", "CA-ROLES", "CA-TENANTS", "CA-PWPOLICY")
buildonly_cit = by_slug(bo, "CIT-GRIEVANCES")

screens = (glob + login_scr + citizen_journey + buildonly_cit
           + eo + dm + ca + buildonly_admin + parity_scr)

master = {"portal": "National Portal for Transgender Persons", "idPrefix": "TG",
          "generated": "2026-07-10",
          "figmaUrl": FURL.format(n="8056-5668"),
          "method": "Design-QC of the TG admin + login screens against the MoSJE Portal Handoff design.",
          "deferred": [
            {"id": "TG-DEBT-01", "title": "Garima Greh login (build) — not captured",
             "reason": "Admin login carries a Garima Greh role tab (design 9379:116062) but no live GG account was provided."},
            {"id": "TG-DEBT-02", "title": "Login states designed-but-not-built",
             "reason": "Registration, credential-recovery, OTP-flow and mobile login states exist in Figma only (no build to compare)."}],
          "screens": screens}
os.makedirs(DEST, exist_ok=True)
json.dump(master, open(os.path.join(OUT, "audit-master-final.json"), "w"), indent=2, ensure_ascii=False)
json.dump(master, open(os.path.join(DEST, "audit-master.json"), "w"), indent=2, ensure_ascii=False)

n = sum(len(s["findings"]) for s in screens)
print(f"screens={len(screens)} findings={n} pin-failures={len(G.FAILURES)}")
r = subprocess.run(["python3", os.path.join(DEST, "generate_pdf.py")], capture_output=True, text=True, timeout=900)
print(r.stdout[-400:])
if r.returncode != 0: print("GEN ERR:", r.stderr[-900:])

# ---------- MASTER tracker: TG sheet + Coverage – TG + Rollup row (matches the NHAPOA format) ----------
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
MASTER = "/Users/akashk/Documents/Projects/MoSJE/docs/qc/MoSJE-Portal-QC-Tracker.xlsx"
SHEET = "TG"; COVSHEET = "Coverage – TG"
COLS = ["ID", "Screen", "Category", "Severity", "Issue (Design → Built)", "Recommended Fix (Dev)",
        "Figma URL", "Live URL", "Status", "Assignee", "Date", "Notes", "Scope"]
WIDTHS = [18, 30, 20, 11, 46, 42, 30, 30, 12, 13, 12, 30, 10]
NAVY = PatternFill("solid", fgColor="003366"); HF = Font(name="Noto Sans", bold=True, color="FFFFFF", size=11)
ZEBRA = PatternFill("solid", fgColor="F4F7FB")
wb = openpyxl.load_workbook(MASTER)
for nm in (SHEET, COVSHEET):
    if nm in wb.sheetnames: del wb[nm]
idx = wb.sheetnames.index("Coverage – NHAPOA") + 1 if "Coverage – NHAPOA" in wb.sheetnames else len(wb.sheetnames)
ws = wb.create_sheet(SHEET, index=idx)
ws.append(COLS)
for c in ws[1]: c.fill = NAVY; c.font = HF; c.alignment = Alignment(vertical="center")
rows = 0
for sc in screens:
    for f in sc["findings"]:
        rows += 1
        issue = f'Figma: {f["figma"]}  →  Live: {f["live"]}'
        notes = f"env: {sc.get('env','dev')}" + ("  ·  " + sc["note"] if sc.get("note") else "")
        figu = sc.get("figmaUrl") or ""; livu = sc.get("liveUrl") or ""
        ws.append([f["id"], sc["name"], f["axis"], f["severity"], issue, f["fix"], figu, livu,
                   "Open", "", "", notes, f.get("scope", "Screen")])
        r_ = ws.max_row
        for c in ws[r_]:
            c.font = Font(name="Noto Sans", size=10); c.alignment = Alignment(vertical="top", wrap_text=True)
            if r_ % 2 == 0: c.fill = ZEBRA
        if figu: ws.cell(r_, 7).hyperlink = figu; ws.cell(r_, 7).font = Font(name="Noto Sans", size=10, color="1558B0", underline="single")
        if livu: ws.cell(r_, 8).hyperlink = livu; ws.cell(r_, 8).font = Font(name="Noto Sans", size=10, color="1558B0", underline="single")
for i, w in enumerate(WIDTHS, 1): ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
ws.freeze_panes = "A2"; ws.auto_filter.ref = f"A1:M{ws.max_row}"
last = max(ws.max_row, 400)
for form, rng in [('"Blocker,Major,Minor,Nit"', f"D2:D{last}"),
                  ('"Layout & Spacing,Color & Token,Typography,Components & States,Content & Iconography,Responsive & A11y"', f"C2:C{last}"),
                  ('"Open,In Progress,Fixed,Won\'t Fix,Verified"', f"I2:I{last}"),
                  ('"Global,Screen"', f"M2:M{last}")]:
    dv = DataValidation(type="list", formula1=form); ws.add_data_validation(dv); dv.add(rng)

cov_ws = wb.create_sheet(COVSHEET, index=wb.sheetnames.index(SHEET) + 1)
cov_ws.append(["Screen", "Section", "Figma URL", "Has Design?", "Built?", "QC Status", "# Findings", "Notes"])
for c in cov_ws[1]: c.fill = NAVY; c.font = HF; c.alignment = Alignment(vertical="center")
for sc in screens:
    if sc.get("_role") == "Global": continue
    status = "Reviewed" if sc["findings"] else ("Coverage — pending pins" if sc.get("figmaImg") else "Build-only")
    cov_ws.append([sc["name"], sc.get("_role", ""), sc.get("figmaUrl") or "",
                   "Yes" if sc.get("figmaImg") else "No", "Yes", status,
                   f"=COUNTIF('{SHEET}'!$B:$B,\"{sc['name']}\")", sc.get("note", "")])
    r_ = cov_ws.max_row
    for c in cov_ws[r_]:
        c.font = Font(name="Noto Sans", size=10); c.alignment = Alignment(vertical="top", wrap_text=True)
        if r_ % 2 == 0: c.fill = ZEBRA
for i, w in enumerate([42, 22, 46, 12, 10, 20, 11, 34], 1): cov_ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
cov_ws.freeze_panes = "A2"; cov_ws.auto_filter.ref = f"A1:H{cov_ws.max_row}"

# Rollup row for TG
rs = wb["Rollup"]; hdr = None
for rr in range(1, rs.max_row + 1):
    if rs.cell(rr, 1).value == "Portal": hdr = rr; break
tg_row = None; empty = None
for rr in range((hdr or 4) + 1, rs.max_row + 3):
    v = rs.cell(rr, 1).value
    if v == SHEET: tg_row = rr; break
    if v in (None, "") and empty is None: empty = rr
target = tg_row or empty or (rs.max_row + 1)
rs.cell(target, 1, SHEET); rs.cell(target, 2, f"=COUNTA('{SHEET}'!$A$2:$A$400)")
for ci, (col, val) in enumerate([("D", "Blocker"), ("D", "Major"), ("D", "Minor"), ("D", "Nit"),
                                 ("I", "Open"), ("I", "Fixed"), ("I", "Verified")], 3):
    rs.cell(target, ci, f"=COUNTIF('{SHEET}'!${col}:${col},\"{val}\")")
for c in rs[target]: c.font = Font(name="Noto Sans", size=11)
wb.save(MASTER)
print(f"TRACKER updated -> {MASTER} · TG rows: {rows} · sheets: {wb.sheetnames}")
