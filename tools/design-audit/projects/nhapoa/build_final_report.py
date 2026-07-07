#!/usr/bin/env python3
"""Single source of truth -> FINAL clean Design-QC PDF + tracker Excel.
Content synced from the Figma review sheet (reviewer points), expanded to
design-intent / live-build / fix, with the correct design frame per screen and
a per-screen env label. Report renders via the clean docs-side generate_pdf.py;
tracker via openpyxl. Both regenerate from the SCREENS list below."""
import json, os, sys, subprocess, re
ENG = os.path.dirname(os.path.abspath(__file__)); TOOL=os.path.dirname(os.path.dirname(ENG))
sys.path.insert(0, os.path.join(TOOL,"engine"))
import qc_geometry as G
PROJ=ENG; OUT=os.path.join(PROJ,"out"); CAP=os.path.join(PROJ,"captures")
DEST="/Users/akashk/Documents/Projects/MoSJE/docs/qc/portals/nhapoa"
FURL="https://www.figma.com/design/gH2vQ62cfg4677YKWuOpLc/MoSJE-Portal--Handoff-?node-id={n}"
def furl(n): return FURL.format(n=n.replace(":","-")) if n else None
def ap(rel): return os.path.join(PROJ,rel) if rel else None

# finding: [sev, axis, title, design-intent, live-build, fix, xpct, ypct, (opt)anchor]
# Global/shared-element issues (accessibility bar, co-branding, sidebar, emblem divider, icon system)
# are consolidated into the GLOBAL section — screens below carry only screen-specific findings.
S=[
 {"slug":"CIT-HOME","role":"Citizen","name":"Home / Dashboard","env":"dev","node":"5986:51876",
  "figma":"captures/figma/CIT-HOME-fresh.png","live":"captures/citizen/hires/CITIZEN-HOME.png","f":[
  ["Minor","Components & States","Hero/action cards style & icons don't match the design","Hero/action cards use the design's icon set, text styles and spacing.","Hero cards render with different icons and typography.","Match the hero-card icon set, text styles and spacing to the design.",50,37],
  ["Minor","Typography","Page-title type style doesn't match the design","Page title uses the design's display type token.","Page title renders in a different type style.","Apply the design's title type token.",34,17],
  ["Minor","Color & Token","Grievance-Closure step subtitle colour off","Step subtitles use the design's muted-text token.","Subtitle colour under the closure steps doesn't match.","Set the closure-step subtitle colour to the design's muted-text token.",50,78],
  ["Nit","Components & States","Masthead 'Admin Login' entry point (good-to-have)","A navy 'Admin Login' button sits top-right for quick admin access.","No Admin Login button in the citizen masthead.","Add the 'Admin Login' button (good-to-have).",90,9],
  ["Nit","Content & Iconography","Footer uses placeholder text","Footer should carry the real ministry footer text.","Footer reuses the design's placeholder text.","Replace with the real production footer text.",50,97],
 ]},
 {"slug":"CIT-RG1","role":"Citizen","name":"Register Grievance · Step 1: Grievance Registration","env":"dev","node":"5986:51959",
  "figma":"captures/figma/CIT-RG1-fresh.png","live":"captures/citizen/hires/CITIZEN-REGISTER-GRIEVANCE.png","f":[
  ["Major","Layout & Spacing","Collapse sidebar by default; cap form width 800px; avoid chatbot overlap","On long multi-step forms the sidebar is collapsed by default and the form column is capped at 800px, keeping CTAs clear of the chatbot.","Sidebar is expanded and the form is full-width, so the chatbot overlaps the CTAs on narrow screens.","Collapse sidebar by default; set form max-width 800px; add bottom spacing equal to the chatbot icon height so it never overlaps CTAs.",50,50],
  ["Minor","Components & States","Active step marker is a hollow ring, not a filled disc","Active step is a solid navy disc with a white numeral.","Active step renders as an outlined ring, reading like inactive steps.","Give the active step the design's solid navy fill + white numeral.",30,17],
  ["Minor","Typography","Upcoming-step label type style off","Upcoming step labels use the design's muted label token.","Upcoming-step labels render in a different style.","Match the upcoming-step label type style.",62,20],
  ["Minor","Components & States","Bottom button should be 'Cancel' and always active","Bottom-left control is 'Cancel', always enabled.","Build shows 'Back', not an always-active cancel.","Relabel to 'Cancel' and keep it active so users can exit the flow.",26,92],
  ["Minor","Layout & Spacing","OTP helper below field; verified/change-number placement per design","'OTP will be sent…' helper sits below the field; the change-number trigger sits beside the input where 'verified' shows, and the verified hint matches the design position.","OTP helper is beside the input; verified/change-number placement differs from design.","Move the OTP helper below the input; place change-number beside the field (where verified shows) and match the verified-hint position to design.",65,86],
  ["Minor","Components & States","'Save & Continue' should enable only when required fields are filled","Primary action enables only when required fields are complete.","'Save and Continue' is enabled even with empty required fields.","Disable it until the step's required fields are valid.",88,92],
 ]},
 {"slug":"CIT-RG2","role":"Citizen","name":"Register Grievance · Step 2: Informer Details","env":"dev","node":"5986:52133",
  "figma":"captures/figma/RG-02-INFORMER.png","live":"captures/citizen/journey/CIT-J-02-step2-informer.png","f":[
  ["Minor","Components & States","Read-only input styling (incl. icon & text) must match","Read-only fields use the design's read-only input style incl. icon + text style.","Read-only inputs render differently.","Apply the design's read-only input style (icon + text) to read-only fields.",60,45],
  ["Minor","Layout & Spacing","Button padding: 16px icon side / 24px other side","Buttons use 16px padding on the icon side, 24px on the other.","Button padding doesn't match 16 / 24px.","Set button padding to 16px (icon side) / 24px (other side).",92,60],
  ["Nit","Components & States","'Use my location' should be a plain text button","'Use my location' is a simple text button.","Rendered as an outlined/box button.","Use the simple text-button style.",92,62],
 ]},
 {"slug":"CIT-RG3","role":"Citizen","name":"Register Grievance · Step 3: Victim Details","env":"dev","node":"5986:52231",
  "figma":"captures/figma/RG-03-VICTIM.png","live":"captures/citizen/journey/CIT-J-03-step3-victim.png","f":[
  ["Minor","Components & States","Aadhaar validation uses a native browser tooltip","Validation uses the app's inline red error, per design.","Wrong Aadhaar shows the browser's native tooltip, inconsistent with the app's inline errors.","Replace native validation with the app's inline red-error styling.",40,58],
  ["Minor","Color & Token","Section-header text colour off","Section headers use the design's header token.","Section-header colour doesn't match.","Set the section-header colour to the design token.",30,30],
  ["Nit","Layout & Spacing","Show the connector line for the ID input section","Design shows a connector line grouping the ID section.","Connector line not shown.","Render the ID-section connector line.",35,54],
 ]},
 {"slug":"CIT-RG4","role":"Citizen","name":"Register Grievance · Step 4: Grievance Details","env":"dev","node":"5986:52298",
  "figma":"captures/figma/RG-04-GRIEVANCE.png","live":"captures/citizen/journey/CIT-J-04-step4-grievance-details.png","f":[
  ["Minor","Components & States","Mark the document upload as mandatory if it blocks submit","Required inputs are marked required per design.","Upload blocks submission but isn't marked mandatory.","If mandatory, mark it required (asterisk + validation copy).",50,80],
  ["Minor","Layout & Spacing","Textarea character count should be right-aligned","Character counter sits bottom-right.","Character count is on the left.","Move the counter to the right.",60,45],
  ["Minor","Components & States","Attached-file card + remove button style must match","Attached-file card and remove button use the design's styling.","They don't match the design.","Match the attached-file card + remove-button style to design.",50,84],
 ]},
 {"slug":"CIT-RG5","role":"Citizen","name":"Register Grievance · Step 5: Review & Submit","env":"dev","node":"5986:52429",
  "figma":"captures/figma/CIT-RG5-REVIEW-fresh.png","live":"captures/citizen/journey/CIT-J-05-step5-review.png","f":[
  ["Minor","Components & States","'Edit details' button style must match the design","The Edit-details button uses the design's button style.","Build's Edit-details button differs.","Match the Edit-details button style to the design.",90,30],
  ["Minor","Components & States","Add a 'View' button for attached documents","The design provides a View action for attached documents.","No View button for the attached document.","Add a 'View' action for the uploaded document.",50,80],
  ["Minor","Color & Token","Consent check stroke too light — looks disabled","The consent checkbox matches the design component (visible check).","The check stroke is so light it looks disabled.","Match the checkbox to the design component so the check reads as active.",8,88],
 ]},
 {"slug":"CIT-RG6","role":"Citizen","name":"Register Grievance · Confirm Submission (modal)","env":"dev","node":"5986:52830",
  "figma":"captures/figma/CIT-RG6-CONFIRM.png","live":"captures/citizen/journey/CIT-J-06-confirm-modal.png","f":[
  ["Minor","Components & States","Confirmation dialog must match the design; drop the Close button","The confirm dialog matches the design; there is no Close (×) — 'Review again' returns to the previous screen.","The build's confirm dialog differs and includes a Close button.","Match the confirm dialog to the design; remove the Close button ('Review again' handles going back).",50,45],
 ]},
 {"slug":"CIT-RG7","role":"Citizen","name":"Register Grievance · Success","env":"dev","node":"5986:52526",
  "figma":"captures/figma/CIT-RG-SUCCESS-fresh.png","live":"captures/citizen/journey/CIT-J-07-success.png","f":[
  ["Minor","Content & Iconography","Check icon used appropriately","Success uses the correct check icon from the icon family.","The success check icon isn't the right one.","Use the correct check icon.",50,22],
 ]},
 {"slug":"CIT-RES","role":"Citizen","name":"Register Rescue","env":"dev","node":None,"undesigned":True,
  "figma":None,"live":"captures/citizen/journey/CIT-RES-01-default.png","f":[
  ["Minor","Typography","Page-title style; no icon in the page title","Title follows the established design language, with no icon.","Title style differs and includes an icon.","Match the page-title style; remove the title icon.",30,17],
  ["Minor","Layout & Spacing","Name / Gender / Mobile in a row; Gender as dropdown","These fields sit in a row; Gender is a dropdown.","Each field is on its own line; Gender isn't a dropdown.","Put Name/Gender/Mobile in a row; make Gender a dropdown.",50,35],
  ["Minor","Layout & Spacing","Location section should follow the established design language","Location title + current-location button follow the established language (no separate divider between sections).","A separate line divider is used between sections.","Follow the established section pattern; drop the extra divider.",50,60],
  ["Minor","Color & Token","OTP modal colours must use the design primary","OTP modal layout/colours use the design's primary colour.","OTP modal uses a random/non-token colour.","Use the design primary colour in the OTP modal (refer to Figma).",50,50],
 ]},
 {"slug":"CIT-TRK1","role":"Citizen","name":"Track Status · Lookup (Reference ID / Mobile)","env":"dev","node":"5986:53203",
  "figma":"captures/figma/CIT-TRK-DEFAULT-design.png","live":"captures/citizen/journey/CIT-TRK-01-default.png","f":[
  ["Minor","Typography","Page title must follow the established design language","Title follows the established language.","Title differs.","Match the page-title style to the design language.",30,17],
  ["Minor","Color & Token","Use the appropriate primary colour from design","Primary elements use the design's primary colour token.","A non-token/random primary colour is used.","Use the design's primary colour token.",50,32],
  ["Minor","Components & States","Card rounded-corner radius must match the design","The lookup card uses the design's corner radius.","Card radius differs from the design.","Match the card corner radius to the design.",50,40],
 ]},
 {"slug":"CIT-TRK3","role":"Citizen","name":"Track Status · Result (View Details)","env":"dev","node":"5986:53285",
  "figma":"captures/figma/CIT-TRK-RESULT-design.png","live":"captures/citizen/journey/CIT-TRK-03-result.png","f":[
  ["Minor","Layout & Spacing","Result layout must match the design","The status/result view layout matches the design (View-Details).","The result layout differs from the design.","Match the result-view layout to the design (ref Figma).",50,40],
  ["Minor","Color & Token","Use the appropriate primary colour from design","Elements use the design's primary colour token.","A non-token/random colour is used.","Use the design's primary colour token.",25,30],
 ]},
 {"slug":"CIT-FAQ","role":"Citizen","name":"Help & FAQs","env":"dev","node":"5986:53669",
  "figma":"captures/figma/CIT-FAQS-fresh.png","live":"captures/citizen/hires/CITIZEN-HELP-FAQS.png","f":[
  ["Major","Responsive & A11y","Accordion opens only from the tiny chevron; whole row should be clickable","Clicking anywhere on the question row expands the answer (a full-width, ≥44px target).","The answer opens only from the small chevron; the question row is a dead element (target far below the 44px minimum).","Make the entire question row the toggle target (keep the chevron as an affordance).",50,40],
  ["Minor","Components & States","Accordion style must match the design","Accordions use the design's style.","Accordion style differs from the design.","Match the accordion style to the design.",50,45],
 ]},
]
# SA_RECUR removed — its recurring patterns are now consolidated in the GLOBAL section (scope=Global).
# grv/usr keep ONLY their screen-specific findings; the portal-wide table/pagination/sidebar/chip/
# page-header/token patterns they exemplified live once in GLOBAL.
SA=[
 {"slug":"SYS-GRV","role":"System Admin","name":"Grievance Monitoring","env":"dev","node":"5986:59480","figma":"captures/figma/SYSTEM-ADMIN-ADMIN-GRIEVANCES.png","live":"captures/live/SYSTEM-ADMIN-ADMIN-GRIEVANCES.png",
  "frame_note":"This list/table view is the reference build for the portal-wide table, pagination, status-chip, sidebar and neutral-token patterns — see the Global Findings section.","f":[
  ["Minor","Components & States","Status badges belong in the Status column with a semantic colour","Status badges sit only in the Status column using the DS semantic colours.","Status badge placement / colour don't use the semantic tokens.","Place status badges in the Status column only and use the appropriate DS semantic colour.",85,25,"Status"],
  ["Minor","Layout & Spacing","Sticky Action column on horizontal scroll","On wide tables the Action column stays pinned to the table edge.","The Action column isn't sticky when the table scrolls horizontally.","Make the Action column sticky to the table's edge on horizontal scroll.",95,25],
 ]},
 {"slug":"SYS-USR","role":"System Admin","name":"User Management","env":"dev","node":"5986:61256","figma":"captures/figma/SYSTEM-ADMIN-ADMIN-USERS.png","live":"captures/live/SYSTEM-ADMIN-ADMIN-USERS.png",
  "frame_note":"Portal-wide table, pagination and neutral-token patterns are consolidated in the Global Findings section; the two items below are specific to this screen.","f":[
  ["Minor","Layout & Spacing","Keep filters on one row when they fit","Filters sit on one row when they fit.","Filter controls wrap to a second row unnecessarily.","Keep the filter controls on one row when they fit; only wrap when needed.",50,20,"Search with name, username, email ID and"],
  ["Minor","Components & States","Consistent 'Clear Filters' button","The clear-filters button is consistent in language, style + position.","Clear-filters button language / style / position is inconsistent.","Use a consistent language, style and position for the clear-filters button.",70,20,"Clear Filters"],
 ]},
]  # SYS-ROL / SYS-REP / SYS-NTF moved to sync_data (sheet-driven detailed findings)

CIT="https://nhapoa-user-dev.mosje.in"; ADM="https://nhapoa-admin-dev.mosje.in"
LIVE_URL={"CIT-HOME":CIT+"/","CIT-RG1":CIT+"/register-grievance","CIT-RG2":CIT+"/register-grievance",
 "CIT-RG3":CIT+"/register-grievance","CIT-RG4":CIT+"/register-grievance","CIT-RG5":CIT+"/register-grievance",
 "CIT-RG6":CIT+"/register-grievance","CIT-RG7":CIT+"/register-grievance","CIT-RES":CIT+"/register-rescue",
 "CIT-TRK1":CIT+"/track-status","CIT-TRK3":CIT+"/track-status","CIT-FAQ":CIT+"/help-faqs",
 "SYS-GRV":ADM+"/admin/grievances","SYS-USR":ADM+"/admin/users","SYS-ROL":ADM+"/admin/roles",
 "SYS-REP":ADM+"/admin/reports","SYS-NTF":ADM+"/admin/notifications","SYS-RECUR":ADM+"/admin/dashboard"}
# Real-geometry pin anchors, keyed by (finding-prefix, ORIGINAL 1-based index in its `f` list).
# Each value is the on-screen text of the element the finding points at; qc_geometry binds the LIVE
# pin to that element's real box from captures/live/<base>.json (largest-font, prefix-forgiving match).
# %-coords stay as a fallback for screens/elements without an extraction row (journey captures, missing
# controls). Omit an entry to keep the hand-tuned % pin for that finding.
ANCHORS={
 "CIT-HOME":{1:"Register Grievance",2:"National Helpline Against Atrocities",
             3:"Grievance Closure Process",5:"Terms & Conditions"},
 "CIT-RG1":{2:"Grievance Registration",3:"Informer Details",4:"Back",5:"OTP will be sent to your registered",
            6:"Save and Continue"},
 "CIT-FAQ":{1:"What is NHAA?",2:"Who can use NHAA?"},
 "CIT-RES":{1:"Register a Rescue",2:"Gender",3:"Location"},
 "CIT-TRK1":{1:"Track Grievance Status",2:"Get OTP & Track Status",3:"Reference ID"},
}
def findings(prefix, flist):
    o=[]
    amap=ANCHORS.get(prefix,{})
    for i,f in enumerate(flist,1):
        sev,axis,title,dz,lz,fix,xp,yp=f[:8]
        anchor=amap.get(i) or (f[8] if len(f)>8 else None)   # on-screen text of the element
        d={"num":i,"id":f"NHA-{prefix}-{i:03d}","element":title,"section":prefix.lower(),"scope":"Screen",
           "axis":canon(axis),"severity":sev,"figma":dz,"live":lz,"fix":fix,"_lpct":(xp,yp),"_fpct":(xp,yp)}
        if anchor: d["_anchor"]=(anchor,0,0)        # binds the LIVE pin to the real element box; %-fallback kept
        o.append(d)
    return o
def screen(s, flist):
    sc={"slug":s["slug"],"name":s["role"]+" — "+s["name"],"env":s.get("env","dev"),
        "figmaImg":ap(s.get("figma")),"liveImg":ap(s["live"]),
        "figmaUrl":furl(s.get("node")),"liveUrl":LIVE_URL.get(s["slug"]),"findings":findings(s["slug"],flist),
        "_role":s["role"],"_node":s.get("node"),"_undesigned":s.get("undesigned",False)}
    if s.get("frame_note"): sc["note"]=s["frame_note"]
    return sc

# ---- synced new-role screens (read back from the Figma 3-column QC sheet) ----
import sync_data as SD
# ---- severity + category recalibration (rubric.md + GIGW/WCAG 2.1 AA) ----
SEV_RANK={"Blocker":0,"Major":1,"Minor":2,"Nit":3}
CATS=("Layout & Spacing","Color & Token","Typography","Components & States","Content & Iconography","Responsive & A11y")
# map any legacy/loose axis onto the 6 canonical categories used by the report + tracker
_CANON={"Accessibility":"Responsive & A11y","Functional":"Components & States",
        "Components & Iconography":"Content & Iconography","Content & iconography":"Content & Iconography"}
def canon(axis): return axis if axis in CATS else _CANON.get(axis,"Components & States")
def classify(text):
    """Rubric-calibrated (severity, category) for a plain-language reviewer bullet."""
    t=text.lower()
    if any(k in t for k in ["accessib","a11y","contrast","toolbar","keyboard","focus-visible","screen reader","wcag","narrow screen","overlap on cta","overlap in narrow"]): cat="Responsive & A11y"
    elif any(k in t for k in ["icon","emoji","logo","co-brand","co branding","material","photo","photograph","blurred","hi-res","wordmark","emblem"," image"]): cat="Content & Iconography"
    elif any(k in t for k in ["colour","color","token","palette","semantic","primary color","primary colour","info color","info colour","status-token","danger","hue"]): cat="Color & Token"
    elif any(k in t for k in ["font","text style","text-style","type style","title style","typography","weight","title case","lowercase","uppercase","transformation","kpi card","card header","header text","header style","tone","copy"]): cat="Typography"
    elif any(k in t for k in ["padding","margin","spacing","align","divider","gap","overflow","width","sticky","radius","corner","hug","full width","full-width","uniform","position","connector","baseline","layout","in a row","1 line","one row","next row"]): cat="Layout & Spacing"
    else: cat="Components & States"
    if any(k in t for k in ["dead element","whole question row","whole row should be clickable","only opens when you click","native tooltip","overflowing outside","overflows outside","overflowing","not rendering","wrong font","serif","blurred","broken"]): sev="Major"
    elif any(k in t for k in ["good to have","good-to-have","extra dot","not needed","isn't required","isn't needed","not required for","not required for 1 page","admin login","placeholder text","realistic footer","use my location","recent reports","connector line","no icon required","icon here is not required","the icon in the page header isn't","1px"]): sev="Nit"
    else: sev="Minor"
    return sev,cat
# lines that are portal-wide repeats (now consolidated in GLOBAL) — drop from the per-screen board
_GLOBAL_SIGS=["recurring admin pattern","shares the do shell","shares the district officer",
  "sho uses the district officer","uses the district officer admin shell","uses the district officer shell",
  "audit against the established","ensure the list, filters, status chips, pagination and table styling",
  "accessibility toolbar reduced","accessibility toolbar is reduced","accessibility bar reduced",
  "the accessibility toolbar","sidebar count badges","pagination is prev / next","prev / next only",
  "notification list-item padding","sla charts should use","sla charts/gauges should use"]
# informational caveats — not findings (become a screen note)
_INFO_SIGS=["empty on dev","no seeded","ignore the red","unexpected error","capture-time",
  "couldn't be verified","could not be verified","no data on dev","review once we have","review once data",
  "no figma frame"]
def _is_global(t):
    tl=t.lower()
    if any(s in tl for s in _GLOBAL_SIGS): return True
    # a short, pure "page header must match…" line — the page-header pattern lives in GLOBAL
    if "page header" in tl and len(tl)<95 and not any(k in tl for k in ["default state","states filter","least privilege","icon here","icon in the page header","permissions","donut","kpi"]):
        return True
    return False
def _is_info(t):
    tl=t.lower().strip("() ")
    return any(s in tl for s in _INFO_SIGS)
def _lead(t,n=74):
    t=t.strip()
    for sep in [" — ",". ",".",";",":"]:
        i=t.find(sep)
        if 14<=i<=n: return t[:i].strip()
    return t[:n].strip()+("…" if len(t)>n else "")
def gen(prefix, issues, undesigned):
    """Turn a reviewer's plain-language bullets into screen findings, dropping portal-wide repeats
    (now in GLOBAL) and informational caveats. Returns (findings, has_global, has_info)."""
    out=[]; i=0; has_global=False; has_info=False
    fig_intent=("Match the established admin design language." if undesigned
                else "See the Figma design frame ↗ (linked on the board) for the intended design.")
    for line in issues.split("\n"):
        c=line.strip().lstrip("•*- ").strip()
        if not c: continue
        if _is_info(c): has_info=True; continue
        if _is_global(c): has_global=True; continue
        i+=1; sev,ax=classify(c)
        out.append({"num":i,"id":f"NHA-{prefix}-{i:03d}","element":_lead(c),"section":prefix.lower(),
                    "scope":"Screen","axis":ax,"severity":sev,"figma":fig_intent,"live":c,"fix":c})
    return out, has_global, has_info
# ---- element-anchored markers for full-board screens (place each finding's pin on its real element) ----
_PIN_STOP=set(("the a an and or of to for in on with your by at is are be should must match design style can not "
  "no its it as per within only from into use used using this that these those a− must1 view page all new "
  "case cases list status user detail details should").split())
def _ptoks(s):
    return set(w for w in re.findall(r"[a-z0-9]+",(s or "").lower()) if len(w)>2 and w not in _PIN_STOP)
def _live_rows(liverel):
    if not liverel: return []
    p=os.path.join(CAP,"live",os.path.splitext(os.path.basename(liverel))[0]+".json")
    try: return json.load(open(p)).get("rows",[]) if os.path.exists(p) else []
    except Exception: return []
def _pinpos(r,Hl):
    x=r["x"]+(110 if r.get("w",0)>700 else r.get("w",0)/2); y=r["y"]+r.get("h",0)/2
    return round(x/1440*100), round(y/max(1,Hl)*100,1)
def _find(rows,*subs):
    for r in sorted(rows,key=lambda r:r.get("y",0)):
        t=(r.get("text") or "").lower()
        if r.get("w") and any(s in t for s in subs): return r
    return None
def _pin_for(bullet,rows,Hl):
    b=bullet.lower(); bt=_ptoks(bullet)
    best=None;bs=0
    for r in rows:
        if not r.get("text") or not r.get("w"): continue
        sh=len(bt & _ptoks(r["text"]))
        if sh>bs or (sh==bs and sh>0 and (r.get("fontSize",0)>(best.get("fontSize",0) if best else 0))): bs=sh;best=r
    if bs>=2 or (bs==1 and best and best.get("fontSize",0)>=15): return _pinpos(best,Hl)   # confident text match
    def h1():
        c=[r for r in rows if r.get("w") and 120<=(r.get("y") or 0)<=300]
        return max(c,key=lambda r:r.get("fontSize",0)) if c else None
    if any(k in b for k in ["accessib","a11y","toolbar","contrast"," widget"]): return 88,round(2000/max(1,Hl),1)
    if any(k in b for k in ["pagination","prev / next","prev/next","page-size","rows-per-page","rows per page","numbered page","long scroll"]): return 50,96
    if any(k in b for k in ["sidebar","count badge"]): r=_find(rows,"dashboard","grievance monitoring"); return _pinpos(r,Hl) if r else (11,20)
    if any(k in b for k in ["page header","header style","page title"]) or b.startswith("page header") or b.startswith("the page header"): r=h1(); return _pinpos(r,Hl) if r else (30,12)
    if "kpi" in b:
        cands=[r for r in rows if r.get("w") and (r.get("fontSize") or 0)>=26 and (r.get("y") or 0)>=250]
        r=min(cands,key=lambda r:r.get("y",0)) if cands else _find(rows,"total grievances","cases","compliance")
        return _pinpos(r,Hl) if r else (30,round(31000/max(1,Hl),1))
    if any(k in b for k in ["chart","legend","graph","funnel","trend","donut","distribution","bar "]): r=_find(rows,"trend","funnel","chart","distribution","monthly","lifecycle"); return _pinpos(r,Hl) if r else (30,round(46000/max(1,Hl),1))
    if any(k in b for k in ["activity feed","feed"]): r=_find(rows,"activity feed","feed"); return _pinpos(r,Hl) if r else (75,round(46000/max(1,Hl),1))
    if "export" in b: r=_find(rows,"export"); return _pinpos(r,Hl) if r else (90,round(16100/max(1,Hl),1))
    if "footer" in b: return 50,97
    return None                                           # no confident element (weak 1-token matches dropped as noise)
def attach_pins(fnd, liverel, figrel, liveH, figH):
    # Markers go on the BUILD board (we have its real element boxes); the design board has no extraction,
    # and build/design page-heights often differ a lot (data-driven), so we do NOT guess a design pin.
    rows=_live_rows(liverel); n=len(fnd)
    for i,f in enumerate(fnd):
        pos=_pin_for(f["fix"],rows,liveH)
        if pos is None:                                   # no confident element -> distribute down the page
            pos=(50, round(8+(i/max(1,n-1) if n>1 else 0.5)*84,1))
        f["livePin"]={"x":pos[0],"y":pos[1]}
        f["figmaPin"]=None

def _mk(slug, role, name, node, liveUrl, figrel, liverel, issues, env="dev"):
    """Build a full-board (no-pin) screen from explicit params — used by both the synced sheet data and
    manually-added screens (inputs/manual-screens.json)."""
    ud = figrel is None; prefix=slug.upper().replace("_","-")
    fnd,has_global,has_info=gen(prefix, issues, ud)
    liveH=G.imgdims(liverel,PROJ)[1] if liverel else 1000
    figH=G.imgdims(figrel,PROJ)[1] if figrel else liveH
    fbox=[0,0,1440,figH]; lbox=[0,0,1440,liveH]
    for f in fnd:
        f["figmaBox"]=fbox; f["liveBox"]=lbox; f["sectionBox"]=lbox
    attach_pins(fnd, liverel, figrel, liveH, figH)   # place a numbered marker on each finding's real element
    note=""
    if has_global: note="Portal-wide patterns on this screen are consolidated in the Global Findings section (each tagged Scope: Global)."
    if has_info: note=(note+" ").strip()+" Some list rows / sub-states couldn't be verified on this dev build (no seeded data)."
    sc={"slug":prefix,"name":name,"env":env,"figmaImg":ap(figrel) if figrel else None,"liveImg":ap(liverel),
        "figmaUrl":furl(node) if node else None,"liveUrl":liveUrl,"findings":fnd,
        "_role":role,"_node":node,"_undesigned":ud}
    if note: sc["note"]=note.strip()
    return sc
# screens the reviewer pointed at a REFERENCE design frame (another screen's design reused as the
# visual-language reference) — the mapping gate skips the design↔build title tier for these.
REF_FRAME_SLUGS={"cc-dash","cc-caller","cc-rg","cc-faq","ca-grievances"}
def newscreen(slug):
    role,name,node,liveUrl,figrel,liverel=SD.META[slug]
    sc=_mk(slug, role, name, node, liveUrl, figrel, liverel, SD.ISSUES.get(slug,""))
    if slug in REF_FRAME_SLUGS:
        sc["_refFrame"]=True
        sc["note"]=(sc.get("note","")+" ").strip()+" Design column is a REFERENCE frame (reuse its visual language); not a same-screen mock."
        sc["note"]=sc["note"].strip()
    return sc
def manual_screens():
    """Screens the USER added/corrected in the Figma sheet, fetched into inputs/manual-screens.json:
    [{slug, role, name, node, live, design(rel png|null), build(rel png), issues, env?}].
    Lets a missed screen or a re-mapped pair be folded in with NO code edit — just data + PNGs."""
    mp=os.path.join(PROJ,"inputs","manual-screens.json")
    out=[]
    if os.path.exists(mp):
        for e in json.load(open(mp)):
            out.append(_mk(e["slug"], e.get("role","Added"), e["name"], e.get("node"), e.get("live"),
                           e.get("design"), e["build"], e.get("issues",""), e.get("env","dev")))
    return out

# ---- GLOBAL section: shared elements that repeat across citizen + admin (fix once) ----
# Each global finding gets its OWN DESIGN|BUILD board pointing at a REPRESENTATIVE design frame
# (with its Figma-frame ↗ link + live route) so a dev can open the source-of-truth frame and see
# the pattern — the fix text names that it applies to every screen with the element.
# ref -> (figma node, design png, build png, live route, label)
GREFS={
 "HOME":   ("5986:51876","captures/figma/CIT-HOME-fresh.png","captures/citizen/hires/CITIZEN-HOME.png",CIT+"/","Citizen — Home"),
 "GRV":    ("5986:59480","captures/figma/SYSTEM-ADMIN-ADMIN-GRIEVANCES.png","captures/live/SYSTEM-ADMIN-ADMIN-GRIEVANCES.png",ADM+"/admin/grievances","System Admin — Grievance Monitoring"),
 "SYSDASH":("5986:59234","captures/figma/roles/SYS/dash.png","captures/live/SYSTEM-ADMIN-ADMIN-DASHBOARD.png",ADM+"/admin/dashboard","System Admin — Dashboard"),
 "NTF":    ("5986:61146","captures/figma/SYSTEM-ADMIN-ADMIN-NOTIFICATIONS.png","captures/live/SYSTEM-ADMIN-ADMIN-NOTIFICATIONS.png",ADM+"/admin/notifications","System Admin — Notifications"),
}
# (severity, category, ref, title, design-intent, live-build, fix, pinx%, piny%)  — severity-ordered
GLOBAL=[
 ("Blocker","Responsive & A11y","HOME","Accessibility toolbar missing from the government masthead",
  "The gov masthead carries the full inline accessibility toolset (A−/A/A+, contrast) plus an accessibility icon that opens the UX4G Accessibility Widget — a GIGW/UX4G-mandated control set present on every page.",
  "On every screen (citizen + admin) the inline toolset is reduced to a single glyph and the UX4G widget is a detached floating FAB, so the mandated accessibility controls are absent from the masthead.",
  "Restore the inline A−/A/A+ + contrast controls in the gov-bar on every page and trigger the UX4G widget from the bar's accessibility icon (not a floating FAB). Applies to every screen (citizen + admin).",86,3),
 ("Major","Content & Iconography","HOME","Co-branding logos + Digital India wordmark mis-render (font-loading)",
  "Co-branding lockups (National Emblem, Digital India, SAMAVESH) render crisply from correct hi-res assets in the correct typeface.",
  "The Digital India logo renders incorrectly and its wordmark falls back to a serif face (webfont not loading); the SAMAVESH login-hero lockup is blurry. Repeats on every masthead.",
  "Use correct hi-res/SVG co-branding assets and fix the wordmark font-loading fallback. Applies to the masthead on every screen and the login hero.",72,8),
 ("Major","Color & Token","GRV","Text off the neutral tokens (headings + muted)",
  "Dark text uses Text/Primary #003366 / Text/Dark #1f2937; muted/secondary text uses Text/Hint #374151.",
  "Headings render near-black slate (#0f172a / #111827) and muted labels render gray-500/400 (#6b7280 / #9ca3af), flattening the colour hierarchy on every admin screen.",
  "Map headings and muted/secondary text to the neutral text tokens. Applies to every admin screen (shown here on the Grievance-Monitoring list).",30,42),
 ("Major","Typography","SYSDASH","KPI cards: figures above the type scale, alert accents lost",
  "KPI numerals sit on the display scale (tops at 28/32), KPI card label text matches the design, and status/danger KPI cards keep their accent styling.",
  "KPI numerals render 48–56px, KPI card text styles differ, and alert cards (e.g. SLA Breaches / SLA Compliance) lose their danger/status accent. Repeats on every dashboard.",
  "Bring KPI figures onto the display scale, match KPI card text styles, and restore the status accent on alert KPI cards. Applies to every dashboard.",25,22),
 ("Major","Color & Token","SYSDASH","Chart palette off the status tokens",
  "Data-viz uses the DS status tokens (Info / Warning / Success / Danger) with the brand navy for emphasis.",
  "Charts use ad-hoc hues (e.g. #a66a26 / #3730a3 / #b7131a) instead of the status tokens, and donut segments lack separating gaps. Repeats on every dashboard/analytics screen.",
  "Re-map chart series to the status/brand tokens and add gaps between donut segments. Applies to every dashboard/analytics screen.",38,58),
 ("Minor","Components & States","SYSDASH","Sidebar: no fill, thin divider, consistent icons, count badges",
  "The sidebar sits on the surface with only a 1px divider; nav icons share one stroke width + neutral colour with a filled primary-colour icon for the active item; nav items carry trailing count badges; the collapse control reflects the sidebar state.",
  "The sidebar has a background fill and no divider; icon stroke/colour vary and the active item isn't a filled primary icon; count badges are missing; the collapse button doesn't reflect state. Repeats on every screen with the sidebar.",
  "Remove the sidebar fill and add a 1px divider; normalise icon stroke + neutral colour with a filled primary active icon; add the count-badge component; bind the collapse button to sidebar state. Applies to every screen with the sidebar.",7,30),
 ("Minor","Typography","GRV","Page-header style",
  "The page header uses the design's text style, spacing, action-button and icon.",
  "Page-header text style / spacing / button / icon differ from the design (and some carry an icon the design omits). Repeats on every admin screen.",
  "Match the page-header text style, spacing, button and icon to the design. Applies to every admin screen.",30,12),
 ("Minor","Components & States","GRV","Data tables (header, cells, sticky action, hover)",
  "Table header fill + cell text weight/colour follow the list style; role text is Title Case; on wide tables the Action column stays pinned; row hover uses primary/50.",
  "Cell text weight/colour differ, role text is lowercase, the Action column isn't sticky on horizontal scroll, and row hover uses a neutral tint. Repeats on every list/table view.",
  "Match table header + cell typography, render roles in Title Case, make the Action column sticky on horizontal scroll, and use primary/50 for row hover. Applies to every list/table view.",52,48),
 ("Minor","Components & States","GRV","Pagination",
  "Lists use numbered pages + a rows-per-page selector.",
  "Lists use Prev / Next + 'Page X of N' only. Repeats on every list view.",
  "Adopt numbered pagination + a rows-per-page selector across list views (paginate tables beyond ~10–15 rows; hide it for a single page). Applies to every list view.",52,92),
 ("Minor","Components & States","GRV","Status chips / badges",
  "Status/SLA pills use the DS semantic colours and sit in the Status column only.",
  "Status/SLA pills don't use the semantic tokens and sometimes appear outside the Status column. Repeats on every list/case view.",
  "Use the DS semantic colours for status/SLA pills and keep them in the Status column. Applies to every list/case view.",80,48),
 ("Minor","Layout & Spacing","SYSDASH","Section / KPI cards",
  "Section and KPI cards share a consistent header text style, uniform heights within a row (width may flex), a divider between a chart and its legend, and charts that fit within the card.",
  "Card header styles differ, row card heights aren't uniform, chart↔legend dividers are missing, and some charts overflow their card. Repeats on every card-based screen.",
  "Standardise card header text, keep uniform card heights per row (flex width), add a chart↔legend divider, and fit charts within the card (horizontal scroll if very wide). Applies to every card-based screen.",68,40),
 ("Minor","Content & Iconography","HOME","Icon system",
  "All icons come from one Material Symbols family at a consistent size/style; no emoji.",
  "Icons mix families/sizes/styles and emoji appear in places (e.g. the citizen track result). Repeats across screens.",
  "Standardise all icons to one Material Symbols family + consistent size/style and replace emoji with Material icons. Applies to every screen.",40,40),
 ("Minor","Components & States","NTF","Notification list items",
  "Notification list-item padding, read/unread styling and icons follow the design; the disabled 'Mark all read' shows only when there are read items.",
  "Notification list-item padding, read/unread styling and icons don't match the design. Repeats on every Notifications screen.",
  "Match notification list-item padding, read/unread styling and icons to the design. Applies to every Notifications screen.",50,34),
 ("Nit","Layout & Spacing","HOME","Emblem–text divider & baseline alignment",
  "The updated design shows the National Emblem beside the ministry text, baseline-aligned, with no divider.",
  "The build shows a divider between the emblem and text and the emblem isn't baseline-aligned. Repeats on every masthead.",
  "Remove the emblem–text divider and align the emblem to the text baseline. Applies to every masthead.",12,8),
]
# STABLE published IDs for the global findings (keyed by title) — frozen so NHA-GLOBAL-0NN always means
# the SAME finding across runs (a dev's tracker status never mis-aligns). Do NOT renumber; only append.
GID={
 "Accessibility toolbar missing from the government masthead":1,
 "Co-branding logos + Digital India wordmark mis-render (font-loading)":2,
 "KPI cards: figures above the type scale, alert accents lost":3,
 "Text off the neutral tokens (headings + muted)":4,
 "Chart palette off the status tokens":5,
 "Page-header style":6,
 "Sidebar: no fill, thin divider, consistent icons, count badges":7,
 "Notification list items":8,
 "Section / KPI cards":9,
 "Icon system":10,
 "Data tables (header, cells, sticky action, hover)":11,
 "Status chips / badges":12,
 "Pagination":13,
 "Emblem–text divider & baseline alignment":14,
}
# Logical grouping of the global findings into a few DESIGN|BUILD pages by their representative frame,
# so the report reads by theme (masthead · dashboards · list views · notifications) instead of one giant
# page or one-page-per-finding. Order the groups by their lowest GID so severity still leads.
GLOBAL_GROUPS=[("HOME","Government masthead & citizen shell"),
               ("SYSDASH","Admin dashboards & cards"),
               ("GRV","Admin list & table views"),
               ("NTF","Notifications")]
def global_screens():
    dim={}
    def _h(rel):
        if rel not in dim: dim[rel]=G.imgdims(rel,PROJ)[1]
        return dim[rel]
    byref={}
    for (sev,cat,ref,title,dz,lz,fix,px,py) in GLOBAL:
        byref.setdefault(ref,[]).append((GID[title],sev,cat,title,dz,lz,fix,px,py))
    out=[]
    for ref,label in GLOBAL_GROUPS:
        items=sorted(byref.get(ref,[]),key=lambda t:t[0])          # by stable GID
        if not items: continue
        node,dpng,bpng,route,rlabel=GREFS[ref]
        Hd=_h(dpng); Hb=_h(bpng)
        fnds=[]
        for (gid,sev,cat,title,dz,lz,fix,px,py) in items:
            # marker number == the finding's stable NHA-GLOBAL id (not a per-page 1,2,3…)
            fnds.append({"num":gid,"id":f"NHA-GLOBAL-{gid:03d}","element":title,"section":"global","scope":"Global",
               "axis":cat,"severity":sev,"figma":dz,"live":lz,"fix":fix,"subO":"Scope: Global — repeats across the portal",
               "figmaBox":[0,0,1440,Hd],"liveBox":[0,0,1440,Hb],"sectionBox":[0,0,1440,Hb],
               "figmaPin":{"x":px,"y":py},"livePin":{"x":px,"y":py}})
        out.append({"slug":f"GLOBAL-{ref}","name":f"Global · {label}","env":"dev",
            "figmaImg":ap(dpng),"liveImg":ap(bpng),"figmaUrl":furl(node),"liveUrl":route,
            "findings":fnds,"_role":"Global","_node":node,"_undesigned":False,"_refFrame":True,
            "note":f"Scope: Global — these elements repeat across the portal (citizen + admin); fix once in the design system. The board shows a representative frame ({rlabel}) — open it via 'Figma frame ↗'; each marker's number matches its NHA-GLOBAL id."})
    return out

# ---- ordering pass: severity first (Blocker→Nit) then top-to-bottom; renumber. Multi-board screens
#      also order their boards by severity. Global pages keep their pre-assigned NHA-GLOBAL-0NN ids. ----
def _piny(f):
    p=f.get("livePin") or f.get("figmaPin")
    return p["y"] if p else 50
def order_by_severity(sc):
    if sc.get("_role")=="Global":              # global pages: keep stable num==id-number; order by it
        sc["findings"].sort(key=lambda f:f["num"])
        for f in sc["findings"]: f.setdefault("scope","Global")
        return
    bysec={}
    for f in sc["findings"]: bysec.setdefault(f["section"],[]).append(f)
    for s in bysec: bysec[s].sort(key=lambda f:(SEV_RANK.get(f["severity"],2), _piny(f)))
    secorder=sorted(bysec, key=lambda s:(min(SEV_RANK.get(f["severity"],2) for f in bysec[s]),
                                         min(_piny(f) for f in bysec[s])))
    new=[f for s in secorder for f in bysec[s]]
    for i,f in enumerate(new,1):
        f["num"]=i; f["id"]=f"{f['id'].rsplit('-',1)[0]}-{i:03d}"; f.setdefault("scope","Screen")
    sc["findings"]=new

# pinned screens (Citizen + System Admin) — geometry via qc_geometry.finalize
citizen=[screen(s,s["f"]) for s in S]
sysadmin_designed=[screen(s,s["f"]) for s in SA]
for sc in citizen+sysadmin_designed: G.finalize(sc, eng_dir=CAP, base_dir=PROJ)
G.write_failures(OUT)
globs=global_screens()
# synced new-role screens (full DESIGN|BUILD boards, no pins; boxes set in newscreen)
cit_extra=[newscreen(sl) for sl in SD.CITIZEN_EXTRA_ORDER]
login=[newscreen(sl) for sl in SD.LOGIN_ORDER]
sys_designed_extra=[newscreen(sl) for sl in SD.SYSADMIN_DESIGNED_ORDER]   # rol/rep/ntf/dash/sla/perf/anl/geo
sysu=[newscreen(sl) for sl in SD.SYSU_ORDER]                              # cat/fb (still undesigned)
newroles=[newscreen(sl) for sl in SD.NEWROLE_ORDER]
# report order: GLOBAL first · Citizen · Login · System Admin (pinned grv/usr + designed + undesigned) · admin roles · detail
# + any screens the user added/corrected via the Figma sheet (inputs/manual-screens.json)
screens = globs + citizen + cit_extra + login + sysadmin_designed + sys_designed_extra + sysu + newroles + manual_screens()
for sc in screens: order_by_severity(sc)

master={"portal":"NHAPOA — National Helpline Against Atrocities","idPrefix":"NHA","generated":"2026-07-04",
 "figmaUrl":FURL.format(n="5093-18512"),"method":"","deferred":[],"screens":screens}
json.dump(master, open(os.path.join(OUT,"audit-master.json"),"w"), indent=2, ensure_ascii=False)
json.dump(master, open(os.path.join(DEST,"audit-master.json"),"w"), indent=2, ensure_ascii=False)
# coverage + design↔build mapping sanity (catches missed screens / wrong frame↔capture pairings)
import crosscheck as XC
xc=XC.run(PROJ, OUT, master)
print(f"cross-check gate={xc['gate']} · mismap={xc['stats']['MISMAP']} to-check={xc['stats']['CHECK']} "
      f"· design-titles={xc['stats']['with_design_heading']} (see out/crosscheck.md)")
for row in xc["rows"]:
    if row["verdict"]=="MISMAP": print("  ⚠ MISMAP:", row["screen"], "—", row["why"])
r=subprocess.run(["python3", os.path.join(DEST,"generate_pdf.py")], capture_output=True, text=True, timeout=900)
print(r.stdout[-300:]);
if r.returncode!=0: print("GEN ERR:", r.stderr[-700:])

# ---------- integrate into the MASTER tracker (match the eUtthan Admin sheet format) ----------
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
MASTER="/Users/akashk/Documents/Projects/MoSJE/docs/qc/MoSJE-Portal-QC-Tracker.xlsx"
SHEET="NHAPOA"; COVSHEET="Coverage – NHAPOA"
# canonical category vocabulary of the master tracker
CATMAP={"Accessibility":"Responsive & A11y","Functional":"Components & States","Components & Iconography":"Content & Iconography"}
def cat(a): return CATMAP.get(a,a)
COLS=["ID","Screen","Category","Severity","Issue (Design → Built)","Recommended Fix (Dev)","Figma URL","Live URL","Status","Assignee","Date","Notes","Scope"]
WIDTHS=[16,26,20,11,46,42,30,30,12,13,12,30,10]
NAVY=PatternFill("solid",fgColor="003366"); HF=Font(name="Noto Sans",bold=True,color="FFFFFF",size=11)
ZEBRA=PatternFill("solid",fgColor="F4F7FB")
wb=openpyxl.load_workbook(MASTER)
for nm in (SHEET,COVSHEET):
    if nm in wb.sheetnames: del wb[nm]
ws=wb.create_sheet(SHEET, index=wb.sheetnames.index("eUtthan Admin")+1)
ws.append(COLS)
for c in ws[1]: c.fill=NAVY; c.font=HF; c.alignment=Alignment(vertical="center")
n=0
for sc in screens:
    for f in sc["findings"]:
        n+=1
        issue=f'Figma: {f["figma"]}  →  Live: {f["live"]}'
        notes=f"env: {sc.get('env','dev')}"+("  ·  "+sc["note"] if sc.get("note") else "")
        figu=f.get("figmaUrlO") or sc.get("figmaUrl")   # per-finding override (GLOBAL rows) else screen link
        livu=f.get("liveUrlO") or sc.get("liveUrl")
        ws.append([f["id"], sc["name"], cat(f["axis"]), f["severity"], issue, f["fix"],
                   figu or "", livu or "", "Open", "", "", notes, f.get("scope","Screen")])
        r=ws.max_row
        for c in ws[r]:
            c.font=Font(name="Noto Sans",size=10); c.alignment=Alignment(vertical="top",wrap_text=True)
            if r%2==0: c.fill=ZEBRA
        if figu: ws.cell(r,7).hyperlink=figu; ws.cell(r,7).font=Font(name="Noto Sans",size=10,color="1558B0",underline="single")
        if livu: ws.cell(r,8).hyperlink=livu; ws.cell(r,8).font=Font(name="Noto Sans",size=10,color="1558B0",underline="single")
for i,w in enumerate(WIDTHS,1): ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width=w
ws.freeze_panes="A2"; ws.auto_filter.ref=f"A1:M{ws.max_row}"
last=max(ws.max_row,400)
dv_sev=DataValidation(type="list",formula1='"Blocker,Major,Minor,Nit"'); ws.add_data_validation(dv_sev); dv_sev.add(f"D2:D{last}")
dv_cat=DataValidation(type="list",formula1='"Layout & Spacing,Color & Token,Typography,Components & States,Content & Iconography,Responsive & A11y"'); ws.add_data_validation(dv_cat); dv_cat.add(f"C2:C{last}")
dv_st=DataValidation(type="list",formula1='"Open,In Progress,Fixed,Won\'t Fix,Verified"'); ws.add_data_validation(dv_st); dv_st.add(f"I2:I{last}")
dv_sc=DataValidation(type="list",formula1='"Global,Screen"'); ws.add_data_validation(dv_sc); dv_sc.add(f"M2:M{last}")
# keep eUtthan Admin consistent — drop its Viewport column (col C) once, re-add dropdowns at shifted cols
eu=wb["eUtthan Admin"]
if eu.cell(1,3).value=="Viewport":
    eu.delete_cols(3,1); eu.data_validations.dataValidation=[]
    for form,rng in [('"Blocker,Major,Minor,Nit"',"D2:D400"),
                     ('"Layout & Spacing,Color & Token,Typography,Components & States,Content & Iconography,Responsive & A11y"',"C2:C400"),
                     ('"Open,In Progress,Fixed,Won\'t Fix,Verified"',"I2:I400")]:
        dv=DataValidation(type="list",formula1=form); eu.add_data_validation(dv); dv.add(rng)
    for i,w in enumerate(WIDTHS,1): eu.column_dimensions[openpyxl.utils.get_column_letter(i)].width=w
    eu.auto_filter.ref=f"A1:L{eu.max_row}"; eu.freeze_panes="A2"

# Coverage – NHAPOA
cov=wb.create_sheet(COVSHEET, index=wb.sheetnames.index(SHEET)+1)
cov.append(["Screen","Section","Figma URL","Has Design?","Built?","QC Status","# Findings","Notes"])
for c in cov[1]: c.fill=NAVY; c.font=HF; c.alignment=Alignment(vertical="center")
for sc in screens:
    if sc.get("_role")=="Global": continue   # Global pages are cross-cutting findings, not audited screens
    cov.append([sc["name"], sc["_role"], sc.get("figmaUrl") or "", "No" if sc.get("_undesigned") or not sc["figmaImg"] else "Yes",
                "Yes","In Review", f"=COUNTIF('{SHEET}'!$B:$B,\"{sc['name']}\")", sc.get("note","")])
    r=cov.max_row
    for c in cov[r]:
        c.font=Font(name="Noto Sans",size=10); c.alignment=Alignment(vertical="top",wrap_text=True)
        if r%2==0: c.fill=ZEBRA
for i,w in enumerate([42,14,46,12,10,14,11,34],1): cov.column_dimensions[openpyxl.utils.get_column_letter(i)].width=w
cov.freeze_panes="A2"; cov.auto_filter.ref=f"A1:H{cov.max_row}"

# Rollup: rewrite both portal rows with the shifted columns (severity=D, status=I)
rs=wb["Rollup"]; hdr_row=None
for rr in range(1,rs.max_row+1):
    if rs.cell(rr,1).value=="Portal": hdr_row=rr; break
def rollup_row(rr, sh):
    rs.cell(rr,1,sh); rs.cell(rr,2,f"=COUNTA('{sh}'!$A$2:$A$400)")
    for ci,(col,val) in enumerate([("D","Blocker"),("D","Major"),("D","Minor"),("D","Nit"),("I","Open"),("I","Fixed"),("I","Verified")],3):
        rs.cell(rr,ci,f"=COUNTIF('{sh}'!${col}:${col},\"{val}\")")
    for c in rs[rr]: c.font=Font(name="Noto Sans",size=11)
eu_row=nha_row=empty_row=None
for rr in range((hdr_row or 4)+1, rs.max_row+2):
    v=rs.cell(rr,1).value
    if v=="eUtthan Admin": eu_row=rr
    elif v=="NHAPOA": nha_row=rr
    elif v in (None,"") and empty_row is None: empty_row=rr
if eu_row: rollup_row(eu_row,"eUtthan Admin")
rollup_row(nha_row or empty_row or ((hdr_row or 4)+2), SHEET)

wb.save(MASTER)
# remove the wrong standalone tracker
sa=os.path.join(DEST,"NHAPOA-Design-QC-Tracker.xlsx")
if os.path.exists(sa): os.remove(sa)
print("MASTER TRACKER updated ->", MASTER, "· NHAPOA sheet rows:", n, "· sheets:", wb.sheetnames)
print("findings:", n, "· pin failures:", len(G.FAILURES))
